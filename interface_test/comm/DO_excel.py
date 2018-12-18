import openpyxl,os,json
from openpyxl import load_workbook
from comm.get_project_path import Project_path
from xlrd import book


class Case:
    def __init__(self):
        self.id=None
        self.url=None
        self.method=None
        self.title=None
        self.test_data=None
        self.expect=None
        self.actual=None
        self.book_name=None
        self.sheet_name=None
        
        
class  Do_excel:
    def __init__(self):
        self.excel_data_path=Project_path().project_path()+"test_data"+os.sep
        
    def read_excel(self,book_name,sheet_name):
        read_path=self.excel_data_path+book_name
        wb=load_workbook(read_path)
        sheet=wb[sheet_name]
        rows=sheet.max_row
        cases=[]
        for i in  range(2,rows+1):
            case=Case()
            case.id=sheet.cell(i,1).value
            case.url=sheet.cell(i,2).value
            case.method=sheet.cell(i,3).value
            case.title=sheet.cell(i,4).value
            case.test_data=json.loads(sheet.cell(i,5).value) 
            case.expect=   json.loads(sheet.cell(i,6).value)
            case.actual=sheet.cell(i,7).value
            case.book_name=book_name
            case.sheet_name=sheet_name
            cases.append(case)
        return cases
    
    def write_excel(self,book_name,sheet_name,row,col,value):
        path=self.excel_data_path+book_name
        wb=load_workbook(path)
        sheet=wb[sheet_name]        
        sheet.cell(row,col).value=value
        wb.save(path)

